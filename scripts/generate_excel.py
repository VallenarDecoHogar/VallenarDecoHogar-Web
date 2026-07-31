#!/usr/bin/env python3
"""Generate a professional .xlsx file with all 39 products and their pricing analysis.

Columns:
- ID
- Categoría
- Nombre
- Precio Real (CLP) — Jimmy India
- Markup % aplicado
- Precio Calculado (sin redondeo)
- Precio Final (redondeado hacia arriba)
- Precio Anterior (display, tachado)
- Ganancia por Unidad
- Margen %
- Stock
- Destacado
- URL Jimmy India
"""
import json
import os
import sys

# Try to use openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, NamedStyle
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
    from openpyxl.chart import BarChart, Reference, PieChart
    from openpyxl.chart.label import DataLabelList
except ImportError:
    print("openpyxl no disponible. Instalando...")
    os.system("pip install openpyxl")
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.chart import BarChart, Reference, PieChart
    from openpyxl.chart.label import DataLabelList


# Load data
with open("/tmp/jimmy/excel_data.json") as f:
    products = json.load(f)

# === STYLE TOKENS (matching Vallenar DecoHogar palette) ===
COLOR_CREAM = "FFFBF5"        # background
COLOR_CREAM_BEIGE = "EEE4DB"  # secondary
COLOR_BROWN = "482113"        # primary text
COLOR_BROWN_MED = "986D47"    # secondary text
COLOR_CORAL = "F17B5D"        # accent
COLOR_CORAL_LIGHT = "FF9D73"  # accent light
COLOR_SAGE = "B4BA8F"         # secondary accent
COLOR_TAN = "CEA170"          # tan
COLOR_GREEN = "00BF63"        # success
COLOR_WHITE = "FFFFFF"
COLOR_BORDER = "E5DCC8"

# Fonts
FONT_NAME = "Calibri"
FONT_TITLE = Font(name=FONT_NAME, size=18, bold=True, color=COLOR_BROWN)
FONT_SUBTITLE = Font(name=FONT_NAME, size=11, italic=True, color=COLOR_BROWN_MED)
FONT_HEADER = Font(name=FONT_NAME, size=11, bold=True, color=COLOR_WHITE)
FONT_CELL = Font(name=FONT_NAME, size=10, color=COLOR_BROWN)
FONT_CELL_BOLD = Font(name=FONT_NAME, size=10, bold=True, color=COLOR_BROWN)
FONT_CELL_LINK = Font(name=FONT_NAME, size=10, color=COLOR_CORAL, underline="single")
FONT_TOTAL = Font(name=FONT_NAME, size=11, bold=True, color=COLOR_WHITE)

# Fills
FILL_HEADER = PatternFill("solid", fgColor=COLOR_BROWN)
FILL_SUBHEADER = PatternFill("solid", fgColor=COLOR_CORAL)
FILL_TOTAL = PatternFill("solid", fgColor=COLOR_BROWN_MED)
FILL_ALT_ROW = PatternFill("solid", fgColor=COLOR_CREAM)
FILL_FEATURED = PatternFill("solid", fgColor="FFF4E6")  # light coral for featured
FILL_SECTION = PatternFill("solid", fgColor=COLOR_CREAM_BEIGE)

# Borders
thin_border = Border(
    left=Side(style="thin", color=COLOR_BORDER),
    right=Side(style="thin", color=COLOR_BORDER),
    top=Side(style="thin", color=COLOR_BORDER),
    bottom=Side(style="thin", color=COLOR_BORDER),
)

# Alignments
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")


def create_products_sheet(wb, products):
    """Create the main products sheet with all 39 products."""
    ws = wb.active
    ws.title = "Productos"

    # === Title row ===
    ws.merge_cells("A1:M1")
    cell = ws["A1"]
    cell.value = "Vallenar DecoHogar — Catálogo Completo de Productos"
    cell.font = FONT_TITLE
    cell.fill = PatternFill("solid", fgColor=COLOR_CREAM)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 32

    # Subtitle
    ws.merge_cells("A2:M2")
    cell = ws["A2"]
    cell.value = f"39 productos importados desde Jimmy India · Precios en CLP · Generado el 2026-07-18"
    cell.font = FONT_SUBTITLE
    cell.fill = PatternFill("solid", fgColor=COLOR_CREAM)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 20

    # Empty row
    ws.row_dimensions[3].height = 8

    # === Header row (row 4) ===
    headers = [
        ("ID", 8),
        ("Categoría", 14),
        ("Nombre del Producto", 38),
        ("Precio Real\n(Jimmy India)", 16),
        ("Markup %", 10),
        ("Precio\nCalculado", 14),
        ("Precio Final\n(redondeado)", 16),
        ("Precio Anterior\n(tachado)", 16),
        ("Ganancia\npor Unidad", 14),
        ("Margen %", 11),
        ("Stock", 9),
        ("Destacado", 11),
        ("URL Jimmy India", 38),
    ]
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[4].height = 36

    # === Data rows ===
    row = 5
    for p in products:
        is_featured = p["featured"] == "Sí"
        is_alt_row = (row - 5) % 2 == 1

        values = [
            p["id"],
            p["category"],
            p["name"],
            p["original_price"],
            f'{p["markup_pct"]}%',
            p["marked_raw"],
            p["final_price"],
            p["old_price"],
            p["profit_per_unit"],
            p["profit_per_unit"] / p["final_price"],
            p["stock"],
            p["featured"],
            "Ver producto en Jimmy India",
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = val
            cell.font = FONT_CELL
            cell.border = thin_border

            # Alternating row colors / featured highlight
            if is_featured:
                cell.fill = FILL_FEATURED
            elif is_alt_row:
                cell.fill = FILL_ALT_ROW

            # Alignment by column
            if col_idx in (1, 5, 11, 12):
                cell.alignment = ALIGN_CENTER
            elif col_idx in (4, 6, 7, 8, 9, 10):
                cell.alignment = ALIGN_RIGHT
            else:
                cell.alignment = ALIGN_LEFT

            # Number formats for currency columns
            if col_idx == 4:  # Precio Real
                cell.number_format = '"$"#,##0'
            elif col_idx == 6:  # Precio Calculado
                cell.number_format = '"$"#,##0'
            elif col_idx == 7:  # Precio Final
                cell.number_format = '"$"#,##0'
                cell.font = FONT_CELL_BOLD
            elif col_idx == 8:  # Precio Anterior
                cell.number_format = '"$"#,##0'
                cell.font = Font(name=FONT_NAME, size=10, color=COLOR_BROWN_MED, strike=True)
            elif col_idx == 9:  # Ganancia
                cell.number_format = '"$"#,##0'
                cell.font = Font(name=FONT_NAME, size=10, bold=True, color=COLOR_GREEN)
            elif col_idx == 10:  # Margen %
                cell.number_format = '0.0%'
                cell.font = Font(name=FONT_NAME, size=10, bold=True, color=COLOR_GREEN)

            # URL link (col 13)
            if col_idx == 13:
                cell.hyperlink = p["supplier_url"]
                cell.font = FONT_CELL_LINK

        ws.row_dimensions[row].height = 22
        row += 1

    # === Totals row ===
    total_row = row
    ws.cell(row=total_row, column=1).value = ""
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
    cell = ws.cell(row=total_row, column=1)
    cell.value = f"TOTAL ({len(products)} productos)"
    cell.font = FONT_TOTAL
    cell.fill = FILL_TOTAL
    cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    cell.border = thin_border

    # Total original price
    total_original = sum(p["original_price"] for p in products)
    cell = ws.cell(row=total_row, column=4)
    cell.value = total_original
    cell.font = FONT_TOTAL
    cell.fill = FILL_TOTAL
    cell.alignment = ALIGN_RIGHT
    cell.border = thin_border
    cell.number_format = '"$"#,##0'

    # Empty markup
    cell = ws.cell(row=total_row, column=5)
    cell.fill = FILL_TOTAL
    cell.border = thin_border

    # Total marked
    total_marked = sum(p["marked_raw"] for p in products)
    cell = ws.cell(row=total_row, column=6)
    cell.value = total_marked
    cell.font = FONT_TOTAL
    cell.fill = FILL_TOTAL
    cell.alignment = ALIGN_RIGHT
    cell.border = thin_border
    cell.number_format = '"$"#,##0'

    # Total final
    total_final = sum(p["final_price"] for p in products)
    cell = ws.cell(row=total_row, column=7)
    cell.value = total_final
    cell.font = FONT_TOTAL
    cell.fill = FILL_TOTAL
    cell.alignment = ALIGN_RIGHT
    cell.border = thin_border
    cell.number_format = '"$"#,##0'

    # Empty
    cell = ws.cell(row=total_row, column=8)
    cell.fill = FILL_TOTAL
    cell.border = thin_border

    # Total profit
    total_profit = sum(p["profit_per_unit"] for p in products)
    cell = ws.cell(row=total_row, column=9)
    cell.value = total_profit
    cell.font = FONT_TOTAL
    cell.fill = FILL_TOTAL
    cell.alignment = ALIGN_RIGHT
    cell.border = thin_border
    cell.number_format = '"$"#,##0'

    # Avg margin
    avg_margin = total_profit / total_final if total_final else 0
    cell = ws.cell(row=total_row, column=10)
    cell.value = avg_margin
    cell.font = FONT_TOTAL
    cell.fill = FILL_TOTAL
    cell.alignment = ALIGN_RIGHT
    cell.border = thin_border
    cell.number_format = '0.0%'

    # Total stock
    total_stock = sum(p["stock"] for p in products)
    cell = ws.cell(row=total_row, column=11)
    cell.value = total_stock
    cell.font = FONT_TOTAL
    cell.fill = FILL_TOTAL
    cell.alignment = ALIGN_CENTER
    cell.border = thin_border

    # Empty
    for col_idx in (12, 13):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.fill = FILL_TOTAL
        cell.border = thin_border

    ws.row_dimensions[total_row].height = 28

    # Freeze panes (below header)
    ws.freeze_panes = "C5"

    # Add auto-filter to header row
    ws.auto_filter.ref = f"A4:M{row - 1}"

    return ws, total_row


def create_summary_sheet(wb, products):
    """Create a summary sheet with stats per category."""
    ws = wb.create_sheet("Resumen por Categoría")

    # Title
    ws.merge_cells("A1:G1")
    cell = ws["A1"]
    cell.value = "Resumen por Categoría"
    cell.font = FONT_TITLE
    cell.fill = PatternFill("solid", fgColor=COLOR_CREAM)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 32

    ws.row_dimensions[2].height = 8

    # Headers
    headers = [
        ("Categoría", 18),
        ("Cantidad", 12),
        ("Precio Real\nPromedio", 18),
        ("Precio Final\nPromedio", 18),
        ("Ganancia\nTotal", 16),
        ("Margen %", 12),
        ("Stock Total", 14),
    ]
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[3].height = 36

    # Group by category
    categories = {}
    for p in products:
        if p["category"] not in categories:
            categories[p["category"]] = []
        categories[p["category"]].append(p)

    row = 4
    for cat, items in categories.items():
        count = len(items)
        avg_real = sum(p["original_price"] for p in items) / count
        avg_final = sum(p["final_price"] for p in items) / count
        total_profit = sum(p["profit_per_unit"] for p in items)
        margin = total_profit / sum(p["final_price"] for p in items)
        stock = sum(p["stock"] for p in items)

        values = [
            cat,
            count,
            avg_real,
            avg_final,
            total_profit,
            margin,
            stock,
        ]

        is_alt_row = (row - 4) % 2 == 1
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = val
            cell.font = FONT_CELL
            cell.border = thin_border
            if is_alt_row:
                cell.fill = FILL_ALT_ROW

            if col_idx == 1:
                cell.alignment = ALIGN_LEFT
                cell.font = FONT_CELL_BOLD
            elif col_idx == 2:
                cell.alignment = ALIGN_CENTER
            elif col_idx in (3, 4, 5):
                cell.alignment = ALIGN_RIGHT
                cell.number_format = '"$"#,##0'
                if col_idx == 5:
                    cell.font = Font(name=FONT_NAME, size=10, bold=True, color=COLOR_GREEN)
            elif col_idx == 6:
                cell.alignment = ALIGN_RIGHT
                cell.number_format = '0.0%'
                cell.font = Font(name=FONT_NAME, size=10, bold=True, color=COLOR_GREEN)
            elif col_idx == 7:
                cell.alignment = ALIGN_CENTER

        ws.row_dimensions[row].height = 26
        row += 1

    # Totals row
    cell = ws.cell(row=row, column=1)
    cell.value = "TOTAL"
    cell.font = FONT_TOTAL
    cell.fill = FILL_TOTAL
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell.border = thin_border

    total_count = len(products)
    cell = ws.cell(row=row, column=2)
    cell.value = total_count
    cell.font = FONT_TOTAL
    cell.fill = FILL_TOTAL
    cell.alignment = ALIGN_CENTER
    cell.border = thin_border

    avg_real_all = sum(p["original_price"] for p in products) / total_count
    cell = ws.cell(row=row, column=3)
    cell.value = avg_real_all
    cell.font = FONT_TOTAL
    cell.fill = FILL_TOTAL
    cell.alignment = ALIGN_RIGHT
    cell.border = thin_border
    cell.number_format = '"$"#,##0'

    avg_final_all = sum(p["final_price"] for p in products) / total_count
    cell = ws.cell(row=row, column=4)
    cell.value = avg_final_all
    cell.font = FONT_TOTAL
    cell.fill = FILL_TOTAL
    cell.alignment = ALIGN_RIGHT
    cell.border = thin_border
    cell.number_format = '"$"#,##0'

    total_profit_all = sum(p["profit_per_unit"] for p in products)
    cell = ws.cell(row=row, column=5)
    cell.value = total_profit_all
    cell.font = FONT_TOTAL
    cell.fill = FILL_TOTAL
    cell.alignment = ALIGN_RIGHT
    cell.border = thin_border
    cell.number_format = '"$"#,##0'

    margin_all = total_profit_all / sum(p["final_price"] for p in products)
    cell = ws.cell(row=row, column=6)
    cell.value = margin_all
    cell.font = FONT_TOTAL
    cell.fill = FILL_TOTAL
    cell.alignment = ALIGN_RIGHT
    cell.border = thin_border
    cell.number_format = '0.0%'

    stock_all = sum(p["stock"] for p in products)
    cell = ws.cell(row=row, column=7)
    cell.value = stock_all
    cell.font = FONT_TOTAL
    cell.fill = FILL_TOTAL
    cell.alignment = ALIGN_CENTER
    cell.border = thin_border

    ws.row_dimensions[row].height = 30

    # Add bar chart for prices by category
    chart = BarChart()
    chart.type = "bar"
    chart.style = 11
    chart.title = "Precio Final Promedio por Categoría"
    chart.y_axis.title = "Categoría"
    chart.x_axis.title = "Precio Final Promedio (CLP)"

    data = Reference(ws, min_col=4, min_row=3, max_row=row-1, max_col=4)
    cats = Reference(ws, min_col=1, min_row=4, max_row=row-1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 18

    ws.add_chart(chart, f"A{row + 3}")

    ws.freeze_panes = "A4"
    return ws


def create_pricing_rules_sheet(wb):
    """Create a sheet explaining the markup rules."""
    ws = wb.create_sheet("Reglas de Precios")

    ws.merge_cells("A1:D1")
    cell = ws["A1"]
    cell.value = "Reglas de Markup y Redondeo"
    cell.font = FONT_TITLE
    cell.fill = PatternFill("solid", fgColor=COLOR_CREAM)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 32

    ws.row_dimensions[2].height = 8

    # Markup rules section
    ws.merge_cells("A3:D3")
    cell = ws["A3"]
    cell.value = "1. Markup aplicado al precio real de Jimmy India"
    cell.font = Font(name=FONT_NAME, size=13, bold=True, color=COLOR_WHITE)
    cell.fill = FILL_SUBHEADER
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[3].height = 26

    headers = ["Condición", "Markup %", "Multiplicador", "Ejemplo"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = thin_border
    ws.row_dimensions[4].height = 28

    rules = [
        ("Precio < $100.000", "150%", "× 2.5", "$990 → $2.475 → $2.500"),
        ("Precio >= $100.000", "75%", "× 1.75", "$100.000 → $175.000"),
    ]
    for i, (cond, markup, mult, example) in enumerate(rules):
        row = 5 + i
        is_alt = i % 2 == 1
        for col_idx, val in enumerate([cond, markup, mult, example], start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = val
            cell.font = FONT_CELL if col_idx > 1 else FONT_CELL_BOLD
            cell.border = thin_border
            if is_alt:
                cell.fill = FILL_ALT_ROW
            cell.alignment = ALIGN_CENTER if col_idx > 1 else ALIGN_LEFT
        ws.row_dimensions[row].height = 24

    # Rounding rules section
    ws.merge_cells(f"A8:D8")
    cell = ws["A8"]
    cell.value = "2. Reglas de redondeo (siempre hacia arriba)"
    cell.font = Font(name=FONT_NAME, size=13, bold=True, color=COLOR_WHITE)
    cell.fill = FILL_SUBHEADER
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[8].height = 26

    headers = ["Rango del Precio Calculado", "Redondear a múltiplos de", "Ejemplo", "Resultado"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=9, column=col_idx)
        cell.value = header
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = thin_border
    ws.row_dimensions[9].height = 28

    rounding_rules = [
        ("Menos de $2.000", "$50", "$1.425 → ceil(28.5) × 50", "$1.450"),
        ("$2.000 – $9.999", "$100", "$2.475 → ceil(24.75) × 100", "$2.500"),
        ("$10.000 – $49.999", "$500", "$12.475 → ceil(24.95) × 500", "$12.500"),
        ("$50.000 – $99.999", "$1.000", "$47.475 → ceil(47.475) × 1000", "$48.000"),
        ("$100.000 o más", "$5.000", "$262.500 → ceil(52.5) × 5000", "$265.000"),
    ]
    for i, (rango, mult, ej, res) in enumerate(rounding_rules):
        row = 10 + i
        is_alt = i % 2 == 1
        for col_idx, val in enumerate([rango, mult, ej, res], start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = val
            cell.font = FONT_CELL if col_idx < 4 else Font(name=FONT_NAME, size=10, bold=True, color=COLOR_CORAL)
            cell.border = thin_border
            if is_alt:
                cell.fill = FILL_ALT_ROW
            cell.alignment = ALIGN_LEFT if col_idx in (1, 3) else ALIGN_CENTER
        ws.row_dimensions[row].height = 22

    # Precio anterior section
    ws.merge_cells("A16:D16")
    cell = ws["A16"]
    cell.value = "3. Precio Anterior (tachado, para mostrar descuento)"
    cell.font = Font(name=FONT_NAME, size=13, bold=True, color=COLOR_WHITE)
    cell.fill = FILL_SUBHEADER
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[16].height = 26

    cell = ws.cell(row=17, column=1)
    cell.value = "Precio Anterior = Precio Final × 1.25, redondeado con las mismas reglas"
    cell.font = FONT_CELL_BOLD
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    cell.border = thin_border
    ws.merge_cells("A17:D17")
    ws.row_dimensions[17].height = 24

    cell = ws.cell(row=18, column=1)
    cell.value = "Ejemplo: Precio Final = $2.500 → Precio Anterior = $3.125 → redondeo → $3.200"
    cell.font = FONT_CELL
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell.border = thin_border
    ws.merge_cells("A18:D18")
    ws.row_dimensions[18].height = 22

    # Variantes section
    ws.merge_cells("A20:D20")
    cell = ws["A20"]
    cell.value = "4. Variantes de producto (multiplicador del precio final)"
    cell.font = Font(name=FONT_NAME, size=13, bold=True, color=COLOR_WHITE)
    cell.fill = FILL_SUBHEADER
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[20].height = 26

    headers = ["Categoría", "Variante", "Multiplicador", "Ejemplo (sobre precio final)"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=21, column=col_idx)
        cell.value = header
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = thin_border
    ws.row_dimensions[21].height = 28

    variants_data = [
        ("Aromaterapia", "15 ml (default)", "× 1.0", "Precio base"),
        ("Aromaterapia", "30 ml", "× 1.8", "+80% sobre precio final"),
        ("Aromaterapia", "50 ml", "× 2.8", "+180% sobre precio final"),
        ("Esoterismo", "Unidad (default)", "× 1.0", "Precio base"),
        ("Esoterismo", "Pack 6 unidades", "× 5.0", "Ahorro vs comprar 6 sueltas"),
        ("Esoterismo", "Pack 12 unidades", "× 9.5", "Mejor precio por unidad"),
        ("Incienso", "Caja 15 gr (default)", "× 1.0", "Precio base"),
        ("Incienso", "Caja 50 gr", "× 3.0", "Mayor cantidad, menor precio por gr"),
        ("Incienso", "Granel 100 gr", "× 5.5", "Mejor para uso intensivo"),
        ("Decoración", "(sin variantes)", "—", "Precio único"),
    ]
    for i, (cat, variant, mult, ej) in enumerate(variants_data):
        row = 22 + i
        is_alt = i % 2 == 1
        for col_idx, val in enumerate([cat, variant, mult, ej], start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = val
            cell.font = FONT_CELL
            cell.border = thin_border
            if is_alt:
                cell.fill = FILL_ALT_ROW
            cell.alignment = ALIGN_LEFT if col_idx in (1, 2, 4) else ALIGN_CENTER
        ws.row_dimensions[row].height = 20

    # Column widths
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 42

    ws.freeze_panes = "A4"

    return ws


def main():
    wb = Workbook()
    wb.properties.creator = "Vallenar DecoHogar"
    wb.properties.title = "Catálogo de Productos con Análisis de Precios"
    wb.properties.subject = "Análisis de markups y márgenes"

    # Sheet 1: All products with full pricing detail
    print("Creating Productos sheet...")
    create_products_sheet(wb, products)

    # Sheet 2: Summary by category
    print("Creating Resumen sheet...")
    create_summary_sheet(wb, products)

    # Sheet 3: Pricing rules explanation
    print("Creating Reglas de Precios sheet...")
    create_pricing_rules_sheet(wb)

    # Save
    output_path = "/home/z/my-project/download/Vallenar-DecoHogar-Catalogo.xlsx"
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)

    file_size = os.path.getsize(output_path) / 1024
    print(f"\n✓ Excel generated: {output_path}")
    print(f"  Size: {file_size:.1f} KB")
    print(f"  Sheets: 3 (Productos, Resumen, Reglas de Precios)")
    print(f"  Products: {len(products)}")


if __name__ == "__main__":
    main()
